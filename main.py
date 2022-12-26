# Read the excel file and convert it to a dataframe
import re
import sys

import openpyxl as op
import numpy as np

# Regex to find the string "X.XXX To X.XXX"
DECIMAL_TO_DECIMAL = r"(\d+\.\d+) To (\d+\.\d+)"

# Regex to find the string "STRING-X.XXX-X.XXX"
STRING_DECIMAL_DECIMAL = r"(\w+)-(\d+\.\d+)-(\d+\.\d+)"


def excel_to_csv(file_path, table):
    wb = op.load_workbook(file_path)
    sheet_names = wb.sheetnames

    print("Total number of sheets: ", len(sheet_names))
    # print("Sheet names: ", sheet_names)
    filtered_list = []
    for sheet_name in sheet_names:
        if re.match(DECIMAL_TO_DECIMAL, sheet_name) or re.match(STRING_DECIMAL_DECIMAL, sheet_name):
            filtered_list.append(wb[sheet_name])

    print("Length of the filtered_list: ", len(filtered_list))
    # only keep 1 sheet for testing
    filtered_list = filtered_list[0:1]
    parse_data(filtered_list, table)


def get_column_index(cell):
    return cell.column


def get_row_index(cell):
    return cell.row


def convert_to_csv(sheet_table, table):
    # Transform the sheet_table and save it to csv
    csv_table = []
    if table == "1":
        header = ["Pointer", "Clarity", "Color", "Price", "Font"]
        # Get the data type of the sheet_table
        for sheet in sheet_table:
            for key, value in sheet.items():
                # print("Key: ", key)
                # print("Value: ", value)
                for item in value:
                    csv_table.append([item["Pointer"], item["Clarity"], item["Color"], item["Price"], item["Font"]])
        save_csv(csv_table, header, table)
    elif table == "2":
        header = ["Pointer", "Clarity", "Cut", "Color", "Florescence", "Font", "Value", "Value_Color"]
        for sheet in sheet_table:
            for key, value in sheet.items():
                # print("Key: ", key)
                # print("Value: ", value)
                for item in value:
                    csv_table.append([item["Pointer"], item["Clarity"], item["Cut"], item["Color"], item["Florescence"],
                                      item["Font"], item["Value"], item["Value_Color"]])
        save_csv(csv_table, header, table)


# Save the csv_table to csv file with table_{table}.csv with header and


def save_csv(csv_table, header, table):
    np.savetxt(f"table_{table}.csv", csv_table, delimiter=",", header=",".join(header), fmt="%s", comments='')
    print("Data inserted successfully")


def parse_table_one(header, pointer, sheet, sheet_items):
    for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
        for cell in row:
            if cell is not None:
                header.append(cell)
        pointer = header[0]
        header = header[1:]
    for row in sheet.iter_rows(min_row=2, values_only=False):
        if row[0].value is None:
            print("Empty row found", row[0].row)
            break
        for cell in row[1:]:
            # skip the empty cells
            if cell.value is None:
                continue
            if cell is not None:
                text = cell
                style = sheet.cell(row=cell.row, column=cell.column).font
                font_style = ""
                if style.b:
                    font_style = "bold".upper()
                elif style.i:
                    font_style = "italic".upper()
                else:
                    font_style = "normal".upper()

                row_dict = {
                    "Pointer": pointer,
                    "Clarity": sheet.cell(row=1, column=cell.column).value,
                    "Color": row[0].value,
                    "Price": text.value,
                    "Font": font_style
                }
                sheet_items.append(row_dict)


def get_clarity_index(cell, clarity_header_row_index, sheet):
    sheet_clarity_header_row = sheet[clarity_header_row_index]  # get the column index of the cell
    cell_column_index = get_column_index(cell) - 1
    # print("Cell column index: ", cell_column_index)
    if sheet_clarity_header_row[cell_column_index].value is not None:
        # print("Cell Value If: ", sheet_clarity_header_row[cell_column_index].value)
        return sheet_clarity_header_row[cell_column_index].value
    else:
        # print("Cell Value Else: ", sheet_clarity_header_row[cell_column_index].value)
        while sheet_clarity_header_row[cell_column_index].value is None:
            cell_column_index -= 1
        return sheet_clarity_header_row[cell_column_index].value


def get_cut_index(cell, cut_header_row_index, sheet):
    sheet_cut_header_row = sheet[cut_header_row_index]  # get the column index of the cell
    cell_column_index = get_column_index(cell) - 1
    return sheet_cut_header_row[cell_column_index].value


def get_florescence_index(cell, florescence_header_index, sheet):
    cell_row_index = get_row_index(cell)
    sheet_florescence_value = sheet.cell(row=cell_row_index, column=florescence_header_index[1]).value
    return sheet_florescence_value


def get_color_index(cell, color_header_index, sheet):
    cell_row_index = get_row_index(cell)
    sheet_color_value = sheet.cell(row=cell_row_index, column=color_header_index[1]).value
    if sheet_color_value is not None:
        return sheet_color_value
    else:
        while sheet_color_value is None:
            cell_row_index -= 1
            sheet_color_value = sheet.cell(row=cell_row_index, column=color_header_index[1]).value
        return sheet_color_value


def get_cell_color(cell):
    argb_color = cell.fill.start_color.index
    if argb_color == "FF000000":
        return "black".upper()
    elif argb_color == "FF00B050":
        return "green".upper()
    elif argb_color == "FFFF0000":
        return "red".upper()
    elif argb_color == "FF0000FF":
        return "blue".upper()
    else:
        return "white".upper()


def get_font_style(cell):
    style = cell.font
    font_style = ""
    if style.b:
        font_style = "bold".upper()
    elif style.i:
        font_style = "italic".upper()
    else:
        font_style = "normal".upper()
    return font_style


def get_pointer_index(cell, pointer_header_index, sheet):
    sheet_pointer_header_row = sheet[pointer_header_index[0]]
    cell_column_index = get_column_index(cell) - 1
    # print("Cell column index: ", cell_column_index)
    if sheet_pointer_header_row[cell_column_index].value is not None:
        # print("Cell Value If: ", sheet_clarity_header_row[cell_column_index].value)
        return sheet_pointer_header_row[cell_column_index].value
    else:
        # print("Cell Value Else: ", sheet_clarity_header_row[cell_column_index].value)
        while sheet_pointer_header_row[cell_column_index].value is None:
            cell_column_index -= 1
        return sheet_pointer_header_row[cell_column_index].value


def parse_table_two(header, pointer, sheet, sheet_items):
    # clarity_header is set of clarity headers
    pointer_header_index = []
    pointer_header = ""
    clarity_header = set()
    clarity_header_row_index = 0
    cut_header = set()
    cut_header_row_index = 0
    color_header_index = []
    florescence_header_index = []
    florescence_header = ["None", "Faint", "Medium", "Strong"]
    # the dictionary to store the table 2 is like this
    # {
    #     "Pointer": 0.20 To 0.229
    #     "Clarity": "IF",
    #     "Cut": "3EX",
    #     "Fluorescence": "None",
    #     "Color": "D",
    #     "Value": "-38",
    #     "Font": "BOLD"
    #     "Value_Color": "Red"
    # }
    for row in sheet.iter_rows(min_row=1, values_only=False):
        # Loop through the first cell of the row is "Range =>"
        if row[0].value == "Range =>":
            for cell in row[1:]:  # skip the first cell of the row and loop through the rest of the cells in the row
                if cell.value is None:
                    continue
                pointer_header_index.append(row[0].row)
                pointer_header = sheet.cell(row=row[0].row, column=2).value
        else:
            pass

    # Loop through the sheet and get the data
    for row in sheet.iter_rows(min_row=pointer_header_index[0], max_row=pointer_header_index[1] - 1, values_only=False):
        if row[0].value == "Clarity =>":
            for cell in row[1:]:
                if cell.value is None:
                    continue
                clarity_header.add(cell.value)
                clarity_header_row_index = row[0].row
        elif row[0].value == "Cut =>":
            for cell in row[1:]:
                if cell.value is None:
                    continue
                cut_header.add(cell.value)
                cut_header_row_index = row[0].row
        elif row[0].value == "Color":
            for cell in row:
                if cell.value is None:
                    continue
                if cell.value == "Color":
                    color_header_index.append(row[0].row)
                    color_header_index.append(cell.column)
                elif cell.value == "Florescence":
                    florescence_header_index.append(row[0].row)
                    florescence_header_index.append(cell.column)
            print("Color header index: ", color_header_index)
            print("Florescence header index: ", florescence_header_index)
        else:
            pass

    for row in sheet.iter_rows(min_row=color_header_index[0] + 1, max_row=pointer_header_index[1] - 1,
                               min_col=3, values_only=False):

        for cell in row:
            pointer_index = get_pointer_index(cell, pointer_header_index, sheet)
            clarity_index = get_clarity_index(cell, clarity_header_row_index, sheet)
            cut_index = get_cut_index(cell, cut_header_row_index, sheet)
            florescence_index = get_florescence_index(cell, florescence_header_index, sheet)
            color_index = get_color_index(cell, color_header_index, sheet)
            cell_color = get_cell_color(cell)
            font_style = get_font_style(cell)

            row_dict = {
                "Pointer": pointer_index,
                "Clarity": clarity_index,
                "Cut": cut_index,
                "Color": color_index,
                "Florescence": florescence_index,
                "Font": font_style,
                "Value": cell.value,
                "Value_Color": cell_color
            }
            print("Row", row_dict)
            sheet_items.append(row_dict)
    # print("Clarity Header: ", clarity_header, clarity_header_row_index)
    # print("Cut Header: ", cut_header, cut_header_row_index)
    print("Pointer Header: ", pointer_header_index)


def parse_data(list, table):
    sheet_table = []
    for sheet in list:
        print("Sheet name: ", sheet.title)
        pointer = []
        header = []
        sheet_items = []

        if table == "1":
            parse_table_one(header, pointer, sheet, sheet_items)
        elif table == "2":
            parse_table_two(header, pointer, sheet, sheet_items)

        sheet_table.append({sheet.title: sheet_items})
    convert_to_csv(sheet_table, table)


if __name__ == '__main__':
    # get file path from command line
    file_path = sys.argv[1]
    table = sys.argv[2]
    # convert excel file to dataframe
    excel_to_csv(file_path, table)

# In the terminal, run the following command
# python main.py "path/to/excel/file" "table_number"
